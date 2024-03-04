# -*- coding: utf-8 -*-
# !/usr/bin/env python
# -------------------------------------------------------------------------------
# Name:         read_excel
# Description:  
# Author:       hulinhui779
# Date:      2024/3/1 16:19
# -------------------------------------------------------------------------------
import openpyxl
import os


class ReadExcelChenShi:
    def __init__(self, path, name=None):
        self.file_path = path
        self.workbook, self.sheet = self.__open_excel(name)

    def __str__(self):
        return f'当前excel信息:总共{self.__excel_max_row()}行{self.__excel_max_cols()}列'

    def __open_excel(self, name):
        if not os.access(self.file_path, os.F_OK):
            raise FileNotFoundError('Excel文件不存在！')
        if not os.access(self.file_path, os.R_OK):
            raise IOError('Excel文件读写错误！')
        workbook = openpyxl.load_workbook(self.file_path)
        if name:
            sheet = workbook[name]
        else:
            sheet = workbook.active  # 获取默认sheet
        return workbook, sheet

    def __excel_max_row(self):
        return self.sheet.max_row

    def __excel_max_cols(self):
        return self.sheet.max_column

    def excel_order_count(self):
        return next(self.sheet.iter_cols(max_col=1, min_row=2, values_only=True))

    def read_data(self, min_row=2):
        data_list = []
        title_list = list(next(self.sheet.iter_rows(max_row=1, values_only=True)))
        for row_value in self.sheet.iter_rows(min_row=min_row, values_only=True):
            if min_row:
                value_list = [value for value in row_value]
                data_list.append(dict(zip(title_list, value_list)))
            else:
                data_list.append(list(row_value))
        self.__colse()
        return data_list

    def __colse(self):
        self.workbook.close()

    if __name__ == '__main__':
        file_path = r'C:\Users\hulinhui779\Downloads\Documents\订单批量出库导入模板.xlsx'
        excel = ReadExcelChenShi(file_path)
        print(excel)
        print(excel.read_data())

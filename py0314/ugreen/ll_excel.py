import re

import openpyxl
from openpyxl.styles import Alignment
import os


class ProbationExcel:
    def __init__(self, path, name=None):
        self.file_path = path
        self.start_row = 5
        self.end_cols = 9
        self.workbook, self.sheet = self.__open_excel(name)

    def __str__(self):
        return f'当前excel信息:总共{self.__excel_max_row()}行{self.__excel_max_cols()}列'

    def __open_excel(self, name):
        if not os.access(self.file_path, os.F_OK):
            raise FileNotFoundError('Excel文件不存在！')
        if not os.access(self.file_path, os.R_OK):
            raise IOError('Excel文件读写错误！')
        workbook = openpyxl.load_workbook(self.file_path)
        if name:
            sheet = workbook[name]
        else:
            sheet = workbook.active  # 获取默认sheet
        return workbook, sheet

    def __excel_max_row(self):
        return self.sheet.max_row

    def __excel_max_cols(self):
        return self.sheet.max_column

    def get_row_data(self, min_row, max_col):
        return list(next(self.sheet.iter_rows(min_row=min_row, max_col=max_col, values_only=True)))

    def user_info_tips(self):
        userinfo_list = [re.sub(r'\s+', '', data) for data in self.get_row_data(3, 9) if data]
        name_text, department_name, *c, employment_date = userinfo_list
        department_name = '部门：' + department_name
        name_text = name_text.strip('部门：')
        print(name_text, department_name, *c, employment_date)

    def write_sheet(self, row, col, value):
        if value:
            self.sheet.cell(row, col).value = value

    @staticmethod
    def read_data(row_data):
        if row_data[1]:
            print(f'{row_data[1]}工作计划如下:')
        print(f'工作内容：{row_data[2]};\n时间段：{row_data[4]}~{row_data[5]};')
        print('-' * 50)

    def write_data(self, rows, cols, is_remark=None):
        try:
            remark = None
            value = int(input('当前任务是否完成？1、是 2、否:'))
            if is_remark:
                remark = input('请输入备注信息！')
            if value == 1:
                self.write_sheet(rows, cols, '是')
                self.write_sheet(rows, cols + 2, remark)
            elif value == 2:
                reason = input('你未完成任务，需要填写未完成的原因：')
                self.write_sheet(rows, cols, '否')
                self.write_sheet(rows, cols + 1, reason)
                self.write_sheet(rows, cols + 2, remark)
            else:
                print('输入有误,跳过不保存')
        except Exception:
            print('非法字符，跳过不保存')

    def add_weekly_task(self):
        row_no = 0
        for index, row_value in enumerate(self.sheet.iter_rows(min_row=self.start_row,
                                                               max_col=self.end_cols,
                                                               values_only=True), 0):
            index += self.start_row
            if any(row_value[2:6]):
                continue
            row_no += index
            break
        week_task_list = self.add_task_list(task_num=4)
        self.insert_data(row_no, week_task_list)
        self.workbook.save(self.file_path)
        self.__close()

    @staticmethod
    def add_task_list(task_num):
        day_num = 1
        week_task_list = []
        while day_num <= task_num:
            week_text = input(f'请输入第几周(格式：第n周)') if day_num == 1 else None
            task_text = input(f'请输入第{day_num}天的工作计划:')
            done_text = input(f'请输入第{day_num}天的达标要求:') if day_num == 1 else None
            start_text = input(f'请输入第{day_num}天的开始时间:')
            end_start = input(f'请输入第{day_num}天的完成时间:')
            week_task_list.append([week_text, task_text, done_text, start_text, end_start])
            day_num += 1
        return week_task_list

    def insert_data(self, row_no, data):
        rows = None
        for data_row, value_list in enumerate(data, 0):  # 数据的行数、数据列表
            rows = row_no + data_row
            for cols, value in enumerate(value_list, 2):  # 第3列开始插入数据
                self.sheet.cell(row=rows, column=cols, value=value)
                self.sheet.cell(row=rows, column=cols).alignment = Alignment(horizontal='center', vertical='center')
            print(f'第{rows}行任务添加成功！')
        self.sheet.merge_cells(f'B{row_no}:B{rows}')
        self.sheet.merge_cells(f'D{row_no}:D{rows}')
        print('本周任务添加成功！')

    def run(self, is_write=False):
        is_have_plan = True
        for index, row_value in enumerate(self.sheet.iter_rows(min_row=self.start_row,
                                                               max_col=self.end_cols,
                                                               values_only=True), 0):
            index += self.start_row
            if not any(row_value[2:6]) or row_value[6]:
                continue
            self.read_data(row_value)
            if is_write:
                self.write_data(index, cols=7)
            is_have_plan = False
        if is_write:
            self.workbook.save(self.file_path)
        if is_have_plan:
            print('太棒了，你的任务都完成了!')
        self.__close()

    def __close(self):
        self.workbook.close()


def main():
    file_path = r'D:\Downloads\document\胡林辉-试用期工作计划表.xlsx'
    excel = ProbationExcel(file_path)
    excel.user_info_tips()
    excel.add_weekly_task()
    excel.run()


if __name__ == '__main__':
    main()
